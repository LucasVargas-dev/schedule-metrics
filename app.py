"""
Gerador de Gantt de Marcos — Engenharia TI
Lê dados dinamicamente de Planning TI.xlsx (abas Planejamento + Marcos)
Gera: Gantt compacto + Conflitos + Dashboard
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta, datetime
import calendar
import unicodedata

# ─── ESTILO ───────────────────────────────────────────────────────────────────
def fill(c):    return PatternFill('solid', start_color=c, end_color=c)
def font(bold=False, sz=9, color='000000', italic=False):
    return Font(name='Arial', bold=bold, size=sz, color=color, italic=italic)
def aln(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def border(*args, left='D0D0D0', right='D0D0D0', top='D0D0D0', bottom='D0D0D0'):
    def s(c): return Side(style='thin', color=c)
    return Border(left=s(left), right=s(right), top=s(top), bottom=s(bottom))

TBORDER = border()

C = {
    'navy':   '1F3864', 'blue':   '2E75B6', 'lblue':  'BDD7EE', 'llblue': 'DEEAF1',
    'green':  '375623', 'lgreen': 'E2EFDA', 'mgreen': '70AD47',
    'amber':  'BF8F00', 'lamber': 'FFF2CC', 'mamber': 'FFD966',
    'red':    'C00000', 'lred':   'FDECEA',
    'purple': '7030A0', 'lpurple':'EAD1F7',
    'gray':   '595959', 'lgray':  'F5F5F5', 'mgray':  'D6DCE4',
    'white':  'FFFFFF', 'dgray':  'A6A6A6', 'black':  '000000',
    'teal':   '006064', 'lteal':  'E0F7FA',
}

# Cores por projeto (bg, fg)
PROJ_COLORS = {
    'Farmabase':        ('E06666', 'FFFFFF'),
    'MES3 Unified':     ('4A86E8', 'FFFFFF'),
    'MES3 - Unified':   ('4A86E8', 'FFFFFF'),
    'MES3 - NOVO':      ('4A86E8', 'FFFFFF'),
    'MES3':             ('1C4587', 'FFFFFF'),
    'MES3 Qualidade':   ('3D5A99', 'FFFFFF'),
    'Hitachi':          ('6AA84F', 'FFFFFF'),
    'Embaquim':         ('7F6000', 'FFFFFF'),
    'ADM':              ('CC0000', 'FFFFFF'),
    'CCPR':             ('FF9900', '000000'),
    'Brasfeed':         ('B45F06', 'FFFFFF'),
    'Alliance One':     ('6D9EEB', 'FFFFFF'),
    'Agraria - Maltaria 2 - MES / MOM': ('D9EAD3', '274E13'),
    'Agraria - Guarapuava': ('D9EAD3', '274E13'),
    'Motasa':           ('D5A6BD', '4A1942'),
    'Agroceres':        ('B6D7A8', '274E13'),
    'Aigle':            ('A4C2F4', '1155CC'),
    'C.Vale F2':        ('92D050', '274E13'),
    'Sooro':            ('FFE599', '7F6000'),
    'Férias':           ('EA4335', 'FFFFFF'),
    'Folga':            ('FBBC04', '000000'),
    'ASO':              ('34A853', 'FFFFFF'),
    'Atestado Médico':  ('FF6D00', 'FFFFFF'),
}

def proj_color(name):
    if not name: return (C['lgray'], C['gray'])
    for key, val in PROJ_COLORS.items():
        if key.lower() in str(name).lower() or str(name).lower() in key.lower():
            return val
    return (C['mgray'], C['black'])

# Cores por tipo de marco (bar color)
TIPO_BG = {
    'Go-live':        'C00000',  # Marco crítico — vermelho
    'Start UP':       'C00000',
    'Testes':         '70AD47',  # Marco normal — verde
    'UAT':            '70AD47',
    'Infraestrutura': '2E75B6',  # azul
    'Demo':           '7030A0',  # roxo
    'Documentação':   'BF8F00',  # âmbar
    'Beta':           '375623',  # verde escuro
    'Desenvolvimento':'2E75B6',  # azul
    'Entrega':        '2E75B6',  # azul
    'Apresentação':   '7030A0',  # roxo
    'Manutenção':     '595959',  # cinza
}

RESTRICTION_TYPES = {
    'Férias', 'Folga', 'Atestado Médico', 'ASO', 'Compromisso',
    'Viagem de retorno', 'Viagem para Lajeado', 'Fechamento de ano',
}

# ─── CARREGA PLANNING TI.XLSX ─────────────────────────────────────────────────
SOURCE_FILE = 'Planning TI.xlsx'
src = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
ws_plan    = src['Planejamento']
ws_marcos  = src['Marcos']

# Mapeamento de datas para colunas na aba Planejamento
# Linha 1: col 2 = 01/12/2025, demais são fórmulas (lidas como None em data_only)
BASE_DATE = date(2025, 12, 1)

def date_to_col_plan(d):
    delta = (d - BASE_DATE).days
    col = delta + 2
    if col < 1 or col > ws_plan.max_column:
        return None
    return col

# ─── MAPEAMENTO PESSOAS → LINHA NA PLANILHA ───────────────────────────────────
NAME_ROW_MAP = {}  # nome_normalizado → row_number
for row in range(1, ws_plan.max_row + 1):
    v = ws_plan.cell(row=row, column=1).value
    if v and isinstance(v, str) and v.strip():
        NAME_ROW_MAP[v.strip().upper()] = row

def _normalize(s):
    """Remove acentos, converte para minúsculas."""
    if not s: return ''
    nfkd = unicodedata.normalize('NFKD', s)
    return ''.join(c for c in nfkd if not unicodedata.combining(c)).lower().strip()

def find_person_row(short_name):
    """Encontra a linha de uma pessoa pelo nome curto (ex: 'Lucas Vargas')."""
    if not short_name or not short_name.strip():
        return None
    norm_short = _normalize(short_name)
    short_parts = set(norm_short.split())

    best_row, best_score = None, 0
    for full_name, row in NAME_ROW_MAP.items():
        norm_full = _normalize(full_name)
        full_parts = set(norm_full.split())
        score = len(short_parts & full_parts)
        if score > best_score:
            best_score = score
            best_row = row
    return best_row if best_score > 0 else None

# Cache de find_person_row para performance
_person_row_cache = {}

def person_row(name):
    if name not in _person_row_cache:
        _person_row_cache[name] = find_person_row(name)
    return _person_row_cache[name]

def get_alloc(person, d):
    """Retorna alocação de uma pessoa numa data."""
    row = person_row(person)
    if row is None: return None
    col = date_to_col_plan(d)
    if col is None: return None
    v = ws_plan.cell(row=row, column=col).value
    return str(v).strip() if v else None

def is_restricted(person, d):
    return get_alloc(person, d) in RESTRICTION_TYPES

# ─── LÊ MARCOS DINAMICAMENTE ─────────────────────────────────────────────────
MILESTONES = []
PROJECTS_ORDER = []  # ordem de aparição dos projetos

for row in range(4, ws_marcos.max_row + 1):
    proj_v = ws_marcos.cell(row=row, column=1).value
    if not proj_v: continue
    proj    = str(proj_v).strip()
    fase    = str(ws_marcos.cell(row=row, column=2).value  or '').strip()
    nome    = str(ws_marcos.cell(row=row, column=3).value  or '').strip()
    data_v  = ws_marcos.cell(row=row, column=4).value
    dur_v   = ws_marcos.cell(row=row, column=5).value
    pres_v  = str(ws_marcos.cell(row=row, column=6).value  or '').strip()
    dev1    = str(ws_marcos.cell(row=row, column=7).value  or '').strip()
    dev2    = str(ws_marcos.cell(row=row, column=8).value  or '').strip()
    rec_v   = str(ws_marcos.cell(row=row, column=9).value  or '').strip()
    status  = str(ws_marcos.cell(row=row, column=10).value or 'Planejado').strip()
    tipo    = str(ws_marcos.cell(row=row, column=12).value or fase).strip()

    if not nome: continue

    if isinstance(data_v, datetime):
        mdate = data_v.date()
    elif isinstance(data_v, date):
        mdate = data_v
    else:
        continue

    dur        = int(dur_v) if dur_v else 1
    presencial = pres_v.lower() in ('sim', 'yes', 's', 'true', '1')
    mandatory  = [d for d in [dev1, dev2] if d]
    rec_list   = [r.strip() for r in rec_v.split(',') if r.strip()] if rec_v else []

    MILESTONES.append({
        'proj': proj, 'fase': fase, 'nome': nome, 'date': mdate,
        'dur': dur, 'presencial': presencial,
        'mandatory': mandatory, 'rec': rec_list,
        'status': status, 'tipo': tipo,
    })

    if proj not in PROJECTS_ORDER:
        PROJECTS_ORDER.append(proj)

# ─── DETECÇÃO DE CONFLITOS ────────────────────────────────────────────────────
def detect_conflicts():
    conflicts = []
    seen = set()
    for ms in MILESTONES:
        for person in ms['mandatory']:
            key = (person, ms['nome'], ms['proj'])
            if key in seen: continue
            for dd in range(ms['dur']):
                check = ms['date'] + timedelta(days=dd)
                if check.weekday() >= 5: continue
                if is_restricted(person, check):
                    alloc = get_alloc(person, check)
                    severity = 'Alta' if ms['tipo'] in ('Go-live','Start UP','UAT','Testes') else 'Média'
                    conflicts.append({
                        'person':    person,
                        'project':   ms['proj'],
                        'milestone': ms['nome'],
                        'date':      ms['date'],
                        'check_date':check,
                        'restriction':alloc,
                        'presencial': ms['presencial'],
                        'severity':  severity,
                    })
                    seen.add(key)
                    break
    return conflicts

CONFLICTS = detect_conflicts()

# ─── CONFIGURAÇÃO DO GANTT ────────────────────────────────────────────────────
today = date.today()

# GANTT_START = segunda-feira da semana atual
# Se fim de semana, avança para próxima segunda
wd = today.weekday()
if wd <= 4:
    GANTT_START = today - timedelta(days=wd)   # segunda desta semana
else:
    GANTT_START = today + timedelta(days=(7 - wd))  # próxima segunda

WEEKS = 16  # semanas exibidas
WEEK_MONDAYS = [GANTT_START + timedelta(weeks=i) for i in range(WEEKS)]

def week_idx(d):
    for i, mon in enumerate(WEEK_MONDAYS):
        if mon <= d < mon + timedelta(days=7):
            return i
    return -1

def weeks_span(d, duration_days):
    weeks = set()
    for dd in range(duration_days):
        wi = week_idx(d + timedelta(days=dd))
        if wi >= 0:
            weeks.add(wi)
    return sorted(weeks)

def ms_end(ms):
    return ms['date'] + timedelta(days=ms['dur'] - 1)

def is_past(ms):
    """Marco completamente anterior à semana atual."""
    return ms_end(ms) < GANTT_START

def short_name(full):
    """'Lucas Pereira Vargas' → 'Lucas V.'"""
    parts = full.strip().split()
    if len(parts) >= 2:
        return f"{parts[0]} {parts[-1][0]}."
    return full

def format_resp(mandatory, rec):
    parts = [short_name(n) for n in mandatory]
    parts += [f"({short_name(n)})" for n in rec]
    return ', '.join(parts)

# ─── CONSTRUÇÃO DO WORKBOOK ───────────────────────────────────────────────────
wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# ABA 1: 📅 Gantt  — layout compacto (Projeto/Marco | Responsáveis | Dur. | semanas)
# ═══════════════════════════════════════════════════════════════════════════════
ws_g = wb.active
ws_g.title = '📅 Gantt'
ws_g.sheet_view.showGridLines = False
ws_g.freeze_panes = 'D4'

LEFT  = 3            # colunas fixas: A, B, C
WCOL  = LEFT + 1     # primeira coluna de semana
TOTAL = LEFT + WEEKS

# ── Linha 1: Título ────────────────────────────────────────────────────────────
ws_g.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL)
c = ws_g.cell(row=1, column=1,
    value=f'GANTT DE MARCOS  —  Engenharia TI  |  Semana {today.strftime("%d/%m/%Y")}')
c.fill = fill(C['navy']); c.font = font(bold=True, sz=12, color=C['white'])
c.alignment = aln('left'); c.border = TBORDER
ws_g.row_dimensions[1].height = 26

# ── Linha 2: Faixas de mês ────────────────────────────────────────────────────
ws_g.merge_cells(start_row=2, start_column=1, end_row=2, end_column=LEFT)
ws_g.cell(row=2, column=1).fill = fill(C['navy'])

cur_month = None
month_start_col = WCOL
for wi, mon in enumerate(WEEK_MONDAYS):
    col = WCOL + wi
    if mon.month != cur_month:
        if cur_month is not None:
            ws_g.merge_cells(start_row=2, start_column=month_start_col, end_row=2, end_column=col - 1)
            mc = ws_g.cell(row=2, column=month_start_col,
                           value=f"{calendar.month_abbr[cur_month].upper()} {WEEK_MONDAYS[wi-1].year}")
            mc.fill = fill(C['blue']); mc.font = font(bold=True, sz=9, color=C['white'])
            mc.alignment = aln('center'); mc.border = TBORDER
        cur_month = mon.month
        month_start_col = col
# Fecha último mês
ws_g.merge_cells(start_row=2, start_column=month_start_col, end_row=2, end_column=WCOL + WEEKS - 1)
mc = ws_g.cell(row=2, column=month_start_col,
               value=f"{calendar.month_abbr[cur_month].upper()} {WEEK_MONDAYS[-1].year}")
mc.fill = fill(C['blue']); mc.font = font(bold=True, sz=9, color=C['white'])
mc.alignment = aln('center'); mc.border = TBORDER
ws_g.row_dimensions[2].height = 14

# ── Linha 3: Cabeçalhos ────────────────────────────────────────────────────────
for ci, lbl in enumerate(['Projeto / Marco', 'Responsáveis', 'Dur.'], 1):
    c = ws_g.cell(row=3, column=ci, value=lbl)
    c.fill = fill(C['navy']); c.font = font(bold=True, sz=9, color=C['white'])
    c.alignment = aln('center'); c.border = TBORDER

for wi, mon in enumerate(WEEK_MONDAYS):
    col = WCOL + wi
    is_now = mon <= today < mon + timedelta(days=7)
    c = ws_g.cell(row=3, column=col, value=mon.strftime('%d/%m'))
    c.fill = fill('4472C4' if is_now else C['navy'])
    c.font = font(bold=True, sz=8, color=C['white'])
    c.alignment = aln('center'); c.border = TBORDER
    ws_g.column_dimensions[get_column_letter(col)].width = 5.5

ws_g.row_dimensions[3].height = 16
ws_g.column_dimensions['A'].width = 30
ws_g.column_dimensions['B'].width = 22
ws_g.column_dimensions['C'].width = 5

# ── Funções de pintura do Gantt ────────────────────────────────────────────────
def paint_bar(ws, row, wis, bg_color, label='', presencial=False):
    """Pinta a barra do marco nas colunas de semana."""
    wis = set(wis)
    for wi in range(WEEKS):
        col = WCOL + wi
        c = ws.cell(row=row, column=col)
        if wi in wis:
            c.fill = fill(bg_color)
            c.font = font(bold=True, sz=7, color='FFFFFF', italic=presencial)
            if wi == min(wis) and label:
                c.value = label[:10]
            c.alignment = aln('center')
            c.border = border(
                top='888888', bottom='888888',
                left='888888' if wi == min(wis) else 'DDDDDD',
                right='888888' if wi == max(wis) else 'DDDDDD',
            )
        else:
            c.fill = fill('F2F2F2' if wi % 2 == 0 else 'FAFAFA')
            c.border = border(left='EEEEEE', right='EEEEEE', top='EEEEEE', bottom='EEEEEE')

def paint_empty_row(ws, row):
    for wi in range(WEEKS):
        col = WCOL + wi
        c = ws.cell(row=row, column=col)
        c.fill = fill('F2F2F2' if wi % 2 == 0 else 'FAFAFA')
        c.border = border(left='EEEEEE', right='EEEEEE', top='EEEEEE', bottom='EEEEEE')

# ── Gera linhas do Gantt ───────────────────────────────────────────────────────
current_row = 4
conflict_set = {(cf['milestone'], cf['project']) for cf in CONFLICTS}

for proj in PROJECTS_ORDER:
    proj_ms = [ms for ms in MILESTONES if ms['proj'] == proj]
    # Só mostra projeto se tiver ao menos um marco presente ou futuro
    if all(is_past(ms) for ms in proj_ms):
        continue

    pb, pf = proj_color(proj)

    # ── Cabeçalho do projeto ──────────────────────────────────────────────────
    ws_g.row_dimensions[current_row].height = 20
    ws_g.merge_cells(start_row=current_row, start_column=1,
                     end_row=current_row, end_column=LEFT)
    c = ws_g.cell(row=current_row, column=1, value=f'  {proj.upper()}')
    c.fill = fill(pb); c.font = font(bold=True, sz=10, color=pf)
    c.alignment = aln('left')
    s = Side(style='thin', color='FFFFFF')
    for wi in range(WEEKS):
        cx = ws_g.cell(row=current_row, column=WCOL + wi)
        cx.fill = fill(pb)
        cx.border = Border(left=s, right=s, top=s, bottom=s)
    current_row += 1

    # ── Linhas de marcos ──────────────────────────────────────────────────────
    for ms in proj_ms:
        past      = is_past(ms)
        tipo      = ms['tipo'] or ms['fase']
        bar_bg    = TIPO_BG.get(tipo, C['blue'])
        has_conf  = (ms['nome'], ms['proj']) in conflict_set
        if has_conf:
            bar_bg = 'C00000'

        resp      = format_resp(ms['mandatory'], ms['rec'])
        dur_lbl   = f"{ms['dur']}d" if ms['dur'] > 1 else '—'
        prefix    = '(P) ' if ms['presencial'] else '+ '
        ms_label  = f"  {prefix}{ms['nome']}"
        wis       = weeks_span(ms['date'], ms['dur'])

        ws_g.row_dimensions[current_row].height = 20

        # Col A: nome do marco
        row_bg = 'F5F5F5' if past else (C['lred'] if has_conf else C['white'])
        row_fg = C['dgray'] if past else (C['red'] if has_conf else C['black'])
        cx = ws_g.cell(row=current_row, column=1, value=ms_label)
        cx.fill = fill(row_bg); cx.border = TBORDER
        cx.font = font(bold=not past, sz=9, color=row_fg, italic=ms['presencial'])
        cx.alignment = aln('left')

        # Col B: responsáveis
        cx = ws_g.cell(row=current_row, column=2, value=resp)
        cx.fill = fill(C['lgray']); cx.border = TBORDER
        cx.font = font(sz=8, color=C['dgray'] if past else C['gray'])
        cx.alignment = aln('left', wrap=True)

        # Col C: duração
        cx = ws_g.cell(row=current_row, column=3, value=dur_lbl)
        cx.fill = fill(C['lgray']); cx.border = TBORDER
        cx.font = font(sz=8, color=C['dgray'] if past else C['gray'])
        cx.alignment = aln('center')

        # Barra de semanas (cinza se passado)
        if wis:
            actual_bg = 'CCCCCC' if past else bar_bg
            paint_bar(ws_g, current_row, wis, actual_bg,
                      label=ms['date'].strftime('%d/%m'),
                      presencial=ms['presencial'])
        else:
            paint_empty_row(ws_g, current_row)

        # Oculta linha se marco está completamente no passado
        if past:
            ws_g.row_dimensions[current_row].hidden = True

        current_row += 1

    # Espaçador entre projetos
    ws_g.row_dimensions[current_row].height = 5
    for col in range(1, TOTAL + 1):
        ws_g.cell(row=current_row, column=col).fill = fill('E0E0E0')
    current_row += 1

# ── Legenda ────────────────────────────────────────────────────────────────────
leg_row = current_row + 1
ws_g.row_dimensions[current_row].height = 8
ws_g.row_dimensions[leg_row].height = 16
ws_g.row_dimensions[leg_row + 1].height = 14

ws_g.merge_cells(start_row=leg_row, start_column=1, end_row=leg_row, end_column=TOTAL)
c = ws_g.cell(row=leg_row, column=1, value=' Legenda:')
c.fill = fill('2D2D2D'); c.font = font(bold=True, sz=8, color='FFFFFF')
c.alignment = aln('left')

leg_items = [
    ('Marco crítico',  'C00000', 'FFFFFF'),
    ('Marco normal',   '70AD47', 'FFFFFF'),
    ('Marco estimado', 'FFD966', '7F6000'),
    ('Marco futuro',   '7030A0', 'FFFFFF'),
    ('(P) Presencial', '444444', 'FFFFFF'),
    ('⚠ Conflito',     'FF6600', 'FFFFFF'),
]
for li, (lbl, lbg, lfg) in enumerate(leg_items):
    col = 1 + li
    if col > TOTAL: break
    c = ws_g.cell(row=leg_row + 1, column=col, value=lbl)
    c.fill = fill(lbg); c.font = font(bold=True, sz=7, color=lfg)
    c.alignment = aln('center'); c.border = TBORDER

# ═══════════════════════════════════════════════════════════════════════════════
# ABA 2: ⚠️ Conflitos
# ═══════════════════════════════════════════════════════════════════════════════
ws_c = wb.create_sheet('⚠️ Conflitos')
ws_c.sheet_view.showGridLines = False
ws_c.freeze_panes = 'A4'

n_conf = len(CONFLICTS)
ws_c.merge_cells('A1:I1')
c = ws_c.cell(row=1, column=1,
    value=f'CONFLITOS DETECTADOS  —  {n_conf} conflito(s) | Cruzamento: Marcos × Alocação Diária')
c.fill = fill(C['red'] if n_conf else C['green'])
c.font = font(bold=True, sz=12, color=C['white'])
c.alignment = aln('left'); c.border = TBORDER
ws_c.row_dimensions[1].height = 26
ws_c.row_dimensions[2].height = 8

CONF_HDRS = ['Dev', 'Projeto', 'Marco', 'Data marco', 'Restrição detectada',
             'Presencial?', 'Severidade', 'Status resolução', 'Substituto / Ação']
CONF_WIDTHS = [22, 16, 30, 12, 24, 11, 12, 18, 28]

for ci, (h, w) in enumerate(zip(CONF_HDRS, CONF_WIDTHS), 1):
    c = ws_c.cell(row=3, column=ci, value=h)
    c.fill = fill(C['red']); c.font = font(bold=True, sz=9, color=C['white'])
    c.alignment = aln('center'); c.border = TBORDER
    ws_c.column_dimensions[get_column_letter(ci)].width = w
ws_c.row_dimensions[3].height = 20

if not CONFLICTS:
    ws_c.merge_cells('A4:I4')
    c = ws_c.cell(row=4, column=1, value='Nenhum conflito detectado.')
    c.fill = fill(C['lgreen']); c.font = font(bold=True, sz=11, color=C['green'])
    c.alignment = aln('center'); c.border = TBORDER
    ws_c.row_dimensions[4].height = 26
else:
    for ri, conf in enumerate(CONFLICTS):
        row = 4 + ri
        ws_c.row_dimensions[row].height = 20
        alt = ri % 2 == 1
        row_bg = 'FFF0F0' if alt else C['lred']
        sev_bg = C['lred']   if conf['severity'] == 'Alta' else C['lamber']
        sev_fg = C['red']    if conf['severity'] == 'Alta' else C['amber']
        vals = [
            conf['person'], conf['project'], conf['milestone'],
            conf['date'], conf['restriction'],
            'Sim' if conf['presencial'] else 'Não',
            conf['severity'], 'Em aberto', '',
        ]
        for ci, v in enumerate(vals, 1):
            bg  = sev_bg if ci == 7 else row_bg
            fg  = sev_fg if ci == 7 else C['black']
            bld = ci in (1, 3, 7)
            fmt = 'DD/MM/YYYY' if ci == 4 else None
            c = ws_c.cell(row=row, column=ci, value=v)
            c.fill = fill(bg); c.font = font(bold=bld, sz=10, color=fg)
            c.alignment = aln('center' if ci in (4, 6, 7, 8) else 'left')
            c.border = TBORDER
            if fmt: c.number_format = fmt

ws_c.auto_filter.ref = f'A3:I{3 + max(1, len(CONFLICTS))}'

# ═══════════════════════════════════════════════════════════════════════════════
# ABA 3: 📊 Dashboard
# ═══════════════════════════════════════════════════════════════════════════════
ws_d = wb.create_sheet('📊 Dashboard')
ws_d.sheet_view.showGridLines = False

ws_d.merge_cells('A1:H1')
c = ws_d.cell(row=1, column=1,
    value='DASHBOARD  —  Visão executiva | Próximos 30 dias | Riscos')
c.fill = fill(C['navy']); c.font = font(bold=True, sz=13, color=C['white'])
c.alignment = aln('left'); c.border = TBORDER
ws_d.row_dimensions[1].height = 26
ws_d.row_dimensions[2].height = 8

# KPIs
ms_30 = sorted(
    [ms for ms in MILESTONES if 0 <= (ms['date'] - today).days <= 30],
    key=lambda x: x['date']
)

KPIS = [
    ('Total marcos',       len(MILESTONES),                            C['blue'],   C['white']),
    ('Próx. 30 dias',      len(ms_30),                                 '0070C0',    C['white']),
    ('Conflitos ativos',   n_conf,                  C['red'] if n_conf else C['green'], C['white']),
    ('Marcos presenciais', sum(1 for m in MILESTONES if m['presencial']), C['purple'], C['white']),
    ('Projetos ativos',    len(PROJECTS_ORDER),                         C['navy'],   C['white']),
    ('Devs monitorados',   len(NAME_ROW_MAP),                           C['teal'],   C['white']),
]

for ki, (lbl, val, kbg, kfg) in enumerate(KPIS):
    col = 1 + ki
    ws_d.column_dimensions[get_column_letter(col)].width = 16
    for r, v, sz, bg, fg, bld in [
        (3, lbl, 9,  kbg,       kfg,    True),
        (4, val, 22, C['white'], kbg,    True),
        (5, '',  8,  C['lgray'], C['gray'], False),
    ]:
        c = ws_d.cell(row=r, column=col, value=v)
        c.fill = fill(bg); c.font = font(bold=bld, sz=sz, color=fg)
        c.alignment = aln('center'); c.border = TBORDER
    ws_d.row_dimensions[3].height = 18
    ws_d.row_dimensions[4].height = 36
    ws_d.row_dimensions[5].height = 10

ws_d.row_dimensions[6].height = 12

# Próximos 30 dias
ws_d.merge_cells('A7:H7')
c = ws_d.cell(row=7, column=1, value='  PRÓXIMOS 30 DIAS  —  Marcos críticos')
c.fill = fill(C['navy']); c.font = font(bold=True, sz=10, color=C['white'])
c.alignment = aln('left'); c.border = TBORDER
ws_d.row_dimensions[7].height = 20

DASH_HDRS = ['Data', 'Projeto', 'Marco', 'Tipo', 'Presencial',
             'Devs obrigatórios', 'Conflito?', 'Dias restantes']
DASH_WIDTHS = [12, 16, 30, 18, 10, 35, 11, 13]
for ci, (h, w) in enumerate(zip(DASH_HDRS, DASH_WIDTHS), 1):
    c = ws_d.cell(row=8, column=ci, value=h)
    c.fill = fill(C['blue']); c.font = font(bold=True, sz=9, color=C['white'])
    c.alignment = aln('center'); c.border = TBORDER
    ws_d.column_dimensions[get_column_letter(ci)].width = w
ws_d.row_dimensions[8].height = 20

if not ms_30:
    ws_d.merge_cells('A9:H9')
    c = ws_d.cell(row=9, column=1, value='Nenhum marco nos próximos 30 dias.')
    c.fill = fill(C['lgreen']); c.font = font(sz=11, color=C['green'])
    c.alignment = aln('center'); c.border = TBORDER
    ws_d.row_dimensions[9].height = 24
else:
    for ri, ms in enumerate(ms_30):
        row = 9 + ri
        ws_d.row_dimensions[row].height = 20
        dias_r   = (ms['date'] - today).days
        has_conf = (ms['nome'], ms['proj']) in conflict_set
        alt      = ri % 2 == 1
        tipo_b   = TIPO_BG.get(ms['tipo'], C['blue'])
        vals = [
            ms['date'], ms['proj'], ms['nome'],
            ms['tipo'], 'Sim' if ms['presencial'] else 'Não',
            ', '.join(ms['mandatory']),
            '⚠ SIM' if has_conf else 'OK', dias_r,
        ]
        for ci, v in enumerate(vals, 1):
            bg = C['llblue'] if alt else C['white']
            fg = C['black']; bld = False; fmt = None
            if ci == 1: fmt = 'DD/MM/YYYY'
            if ci == 4: bg = tipo_b;  fg = C['white']; bld = True
            if ci == 5 and ms['presencial']: bg = C['lpurple']; fg = C['purple']
            if ci == 7:
                bg  = C['lred']   if has_conf else C['lgreen']
                fg  = C['red']    if has_conf else C['green']
                bld = True
            if ci == 8:
                bg  = C['lamber'] if dias_r < 7 else bg
                fg  = C['amber']  if dias_r < 7 else fg
                bld = dias_r < 7
            c = ws_d.cell(row=row, column=ci, value=v)
            c.fill = fill(bg); c.font = font(bold=bld, sz=10, color=fg)
            c.alignment = aln('center' if ci in (1, 4, 5, 7, 8) else 'left', wrap=(ci == 6))
            c.border = TBORDER
            if fmt: c.number_format = fmt

# Risco por projeto
gap = 9 + max(1, len(ms_30)) + 2
ws_d.row_dimensions[gap - 1].height = 12
ws_d.merge_cells(f'A{gap}:H{gap}')
c = ws_d.cell(row=gap, column=1, value='  RISCO POR PROJETO')
c.fill = fill(C['navy']); c.font = font(bold=True, sz=10, color=C['white'])
c.alignment = aln('left'); c.border = TBORDER
ws_d.row_dimensions[gap].height = 20

RISK_HDRS = ['Projeto', 'Marcos', 'Próx. marco', 'Data', 'Conflitos', 'Presenciais', 'Status geral']
for ci, h in enumerate(RISK_HDRS, 1):
    c = ws_d.cell(row=gap + 1, column=ci, value=h)
    c.fill = fill(C['blue']); c.font = font(bold=True, sz=9, color=C['white'])
    c.alignment = aln('center'); c.border = TBORDER
ws_d.row_dimensions[gap + 1].height = 20

for pi, proj in enumerate(PROJECTS_ORDER):
    row = gap + 2 + pi
    ws_d.row_dimensions[row].height = 22
    pms    = [m for m in MILESTONES if m['proj'] == proj]
    future = sorted([m for m in pms if m['date'] >= today], key=lambda x: x['date'])
    next_ms = future[0] if future else None
    n_conf_p = sum(1 for cf in CONFLICTS if cf['project'] == proj)
    n_pres   = sum(1 for m in pms if m['presencial'])
    status   = ('⚠ Em risco' if n_conf_p > 0
                else ('★ Go-live iminente' if any(m['tipo'] in ('Go-live','Start UP') for m in future[:2])
                      else '✓ Normal'))
    s_bg = C['lred'] if n_conf_p > 0 else (C['lamber'] if 'Go-live' in status else C['lgreen'])
    s_fg = C['red']  if n_conf_p > 0 else (C['amber']  if 'Go-live' in status else C['green'])
    pb, pf = proj_color(proj)
    alt = pi % 2 == 1
    vals = [proj, len(pms),
            next_ms['date'] if next_ms else '—',
            n_conf_p, n_pres, status]
    for ci, v in enumerate(vals, 1):
        bg = C['llblue'] if alt else C['white']
        fg = C['black']; bld = False; fmt = None
        if ci == 1: bg = pb; fg = pf; bld = True
        if ci == 3 and isinstance(v, date): fmt = 'DD/MM/YYYY'
        if ci == 4: bg = (C['lred'] if v > 0 else bg); fg = (C['red'] if v > 0 else fg); bld = v > 0
        if ci == 5: bg = (C['lpurple'] if v > 0 else bg); fg = (C['purple'] if v > 0 else fg)
        if ci == 6: bg = s_bg; fg = s_fg; bld = True
        c = ws_d.cell(row=row, column=ci, value=v)
        c.fill = fill(bg); c.font = font(bold=bld, sz=10, color=fg)
        c.alignment = aln('center' if ci in (2, 3, 4, 5, 6) else 'left')
        c.border = TBORDER
        if fmt: c.number_format = fmt

# ─── SALVA ────────────────────────────────────────────────────────────────────
out = 'Cronograma_Gantt.xlsx'
wb.save(out)
sheet_names = [s.encode('ascii', 'replace').decode() for s in wb.sheetnames]
print(f'Salvo: {out}')
print(f'Sheets: {sheet_names}')
print(f'Marcos carregados: {len(MILESTONES)} ({len(PROJECTS_ORDER)} projetos)')
print(f'Conflitos detectados: {n_conf}')
for cf in CONFLICTS:
    d_str = cf['date'].strftime('%d/%m')
    print(f"  ! {cf['person']} - {cf['milestone']} ({d_str}) x {cf['restriction']}")
print(f'GANTT_START: {GANTT_START} (semana atual)')
print(f'Marcos passados (ocultos): {sum(1 for ms in MILESTONES if is_past(ms))}')
print(f'Marcos prox. 30 dias: {len(ms_30)}')
